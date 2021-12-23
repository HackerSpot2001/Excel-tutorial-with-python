#!/usr/bin/python3
from openpyxl import load_workbook

wb_obj = load_workbook("BCA3RD.xlsx")
sheet_obj = wb_obj.active
row = sheet_obj.max_row
column = sheet_obj.max_column
print("\nValue of first column")
emails = []
for i in range(3, row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 5) 
    print(cell_obj.value) 
    emails.append(cell_obj.value)

with open("emails.txt","a+") as f:
    for email in emails:
        f.write(f"{email}\n")